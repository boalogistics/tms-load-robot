from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Border

def set_border(self, ws, cell_range):
     border = Border(left=Side(border_style='thin', color='000000'),
                    right=Side(border_style='thin', color='000000'),
                    top=Side(border_style='thin', color='000000'),
                    bottom=Side(border_style='thin', color='000000'))

    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            cell.border = border

#usage example:
#ws = load_workbook('example.xlsx').get_active_sheet()
#set_broder(ws, "C3:H10")